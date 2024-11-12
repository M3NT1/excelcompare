import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl.styles import PatternFill, Border, Side
import numpy as np


class FileComparisonApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Excel/CSV Összehasonlító")
        self.root.geometry("600x300")

        self.KEY_COLUMN = "Hivatkozási név (kód)"

        # File path variables
        self.file1_path = tk.StringVar()
        self.file2_path = tk.StringVar()

        # Styles
        self.red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
        self.green_fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
        self.purple_fill = PatternFill(start_color='FFB19CD9', end_color='FFB19CD9', fill_type='solid')
        self.thick_border = Border(
            left=Side(style='thick'),
            right=Side(style='thick'),
            top=Side(style='thick'),
            bottom=Side(style='thick')
        )

        # GUI elements
        self.create_widgets()

    def create_widgets(self):
        # File 1 selection
        tk.Label(self.root, text="Első fájl (Forrás):").pack(pady=5)
        tk.Entry(self.root, textvariable=self.file1_path, width=50).pack(pady=5)
        tk.Button(self.root, text="Tallózás", command=lambda: self.browse_file(1)).pack(pady=5)

        # File 2 selection
        tk.Label(self.root, text="Második fájl (Összehasonlítandó):").pack(pady=5)
        tk.Entry(self.root, textvariable=self.file2_path, width=50).pack(pady=5)
        tk.Button(self.root, text="Tallózás", command=lambda: self.browse_file(2)).pack(pady=5)

        # Compare button
        tk.Button(self.root, text="Összehasonlítás", command=self.compare_files).pack(pady=20)

    def browse_file(self, file_num):
        filetypes = [
            ('Excel fájlok', '*.xlsx'),
            ('CSV fájlok', '*.csv'),
            ('Minden fájl', '*.*')
        ]
        filename = filedialog.askopenfilename(
            title=f"Válassza ki a {file_num}. fájlt",
            filetypes=filetypes
        )
        if file_num == 1:
            self.file1_path.set(filename)
        else:
            self.file2_path.set(filename)

    def read_file(self, file_path):
        if file_path.endswith('.csv'):
            return pd.read_csv(file_path)
        else:
            return pd.read_excel(file_path)

    def validate_files(self, df1, df2):
        # Check if key column exists
        if self.KEY_COLUMN not in df1.columns or self.KEY_COLUMN not in df2.columns:
            raise ValueError(f"A '{self.KEY_COLUMN}' oszlop nem található mindkét fájlban!")

        # Check if columns match
        if set(df1.columns) != set(df2.columns):
            raise ValueError("A két excel oszlopainak száma nem egyező, az összehasonlítás nem hajtható végre!")

        return True

    def is_empty_value(self, value):
        """Check if a value is considered empty (None, nan, or empty string)"""
        if pd.isna(value):
            return True
        if isinstance(value, str) and value.strip() == '':
            return True
        return False

    def values_are_equal(self, val1, val2):
        """Compare two values, treating empty values as equal"""
        if self.is_empty_value(val1) and self.is_empty_value(val2):
            return True
        return val1 == val2

    def prepare_side_by_side_comparison(self, df1, df2):
        # Get all unique keys
        all_keys = sorted(set(df1[self.KEY_COLUMN]) | set(df2[self.KEY_COLUMN]))

        # Create empty DataFrames with all keys
        df1_aligned = pd.DataFrame({self.KEY_COLUMN: all_keys})
        df2_aligned = pd.DataFrame({self.KEY_COLUMN: all_keys})

        # Merge with original DataFrames
        df1_aligned = df1_aligned.merge(df1, on=self.KEY_COLUMN, how='left')
        df2_aligned = df2_aligned.merge(df2, on=self.KEY_COLUMN, how='left')

        # Rename columns to distinguish between source and target
        df1_cols = {col: f"{col} (Forrás)" for col in df1_aligned.columns}
        df2_cols = {col: f"{col} (Cél)" for col in df2_aligned.columns}
        df1_aligned = df1_aligned.rename(columns=df1_cols)
        df2_aligned = df2_aligned.rename(columns=df2_cols)

        # Combine the DataFrames side by side
        combined_df = pd.concat([df1_aligned, df2_aligned], axis=1)

        return combined_df, df1_aligned, df2_aligned

    def compare_rows(self, df1, df2):
        stats = {
            'total_rows_source': len(df1),
            'total_rows_target': len(df2),
            'matching_rows': 0,
            'partial_matches': 0,
            'deleted_rows': 0,
            'new_rows': 0,
            'column_differences': {col: 0 for col in df1.columns if col != self.KEY_COLUMN}
        }

        all_keys = set(df1[self.KEY_COLUMN]) | set(df2[self.KEY_COLUMN])

        deleted_rows = []
        new_rows = []
        partial_matches = []
        different_cells = {}  # Dictionary to store different cell positions

        for key in all_keys:
            df1_row = df1[df1[self.KEY_COLUMN] == key]
            df2_row = df2[df2[self.KEY_COLUMN] == key]

            if not df2_row.empty and df1_row.empty:
                new_rows.append(key)
                stats['new_rows'] += 1
            elif not df1_row.empty and df2_row.empty:
                deleted_rows.append(key)
                stats['deleted_rows'] += 1
            elif not df1_row.empty and not df2_row.empty:
                different_columns = []
                row_equal = True

                for col in df1.columns:
                    if col != self.KEY_COLUMN:
                        val1 = df1_row[col].iloc[0]
                        val2 = df2_row[col].iloc[0]

                        if not self.values_are_equal(val1, val2):
                            row_equal = False
                            different_columns.append(col)
                            stats['column_differences'][col] += 1

                if row_equal:
                    stats['matching_rows'] += 1
                else:
                    partial_matches.append(key)
                    different_cells[key] = different_columns
                    stats['partial_matches'] += 1

        return deleted_rows, new_rows, partial_matches, different_cells, stats

    def create_statistics_sheet(self, workbook, stats):
        ws = workbook.create_sheet("Statisztika")

        # Basic statistics
        ws['A1'] = "Összesítő statisztika"
        ws['A3'] = "Forrás fájl sorok száma:"
        ws['B3'] = stats['total_rows_source']
        ws['A4'] = "Cél fájl sorok száma:"
        ws['B4'] = stats['total_rows_target']
        ws['A5'] = "Teljesen egyező sorok:"
        ws['B5'] = stats['matching_rows']
        ws['A6'] = "Részlegesen egyező sorok:"
        ws['B6'] = stats['partial_matches']
        ws['A7'] = "Törölt sorok:"
        ws['B7'] = stats['deleted_rows']
        ws['A8'] = "Új sorok:"
        ws['B8'] = stats['new_rows']

        # Column differences
        ws['A10'] = "Oszloponkénti eltérések"
        row = 11
        for col, count in stats['column_differences'].items():
            ws[f'A{row}'] = col
            ws[f'B{row}'] = count
            row += 1

    def compare_files(self):
        if not self.file1_path.get() or not self.file2_path.get():
            messagebox.showerror("Hiba", "Kérem válassza ki mindkét fájlt!")
            return

        try:
            # Read files
            df1 = self.read_file(self.file1_path.get())
            df2 = self.read_file(self.file2_path.get())

            # Validate files
            self.validate_files(df1, df2)

            # Compare files
            deleted_rows, new_rows, partial_matches, different_cells, stats = self.compare_rows(df1, df2)

            # Prepare side by side comparison
            combined_df, df1_aligned, df2_aligned = self.prepare_side_by_side_comparison(df1, df2)

            # Ask for save location
            save_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel fájlok", "*.xlsx")],
                title="Mentés másként"
            )

            if not save_path:
                return

            # Create Excel writer
            writer = pd.ExcelWriter(save_path, engine='openpyxl')

            # Write combined dataframe
            combined_df.to_excel(writer, sheet_name='Összehasonlítás', index=False)

            # Get workbook and worksheet
            workbook = writer.book
            ws = workbook['Összehasonlítás']

            # Get column indices for both source and target columns
            source_cols = {col: idx for idx, col in enumerate(df1.columns)}
            target_cols = {col: idx + len(df1.columns) for idx, col in enumerate(df2.columns)}

            # Color the rows and mark different cells
            for row_idx, row in enumerate(ws.iter_rows(min_row=2)):
                key = combined_df.iloc[row_idx][f"{self.KEY_COLUMN} (Forrás)"]

                if pd.isna(key):
                    key = combined_df.iloc[row_idx][f"{self.KEY_COLUMN} (Cél)"]
                    # New row - color the right side cells green
                    for cell in row[len(df1.columns):]:
                        cell.fill = self.green_fill
                elif key in deleted_rows:
                    # Deleted row - color the left side cells red
                    for cell in row[:len(df1.columns)]:
                        cell.fill = self.red_fill
                elif key in partial_matches:
                    # Partial match - color all cells light purple
                    for cell in row:
                        cell.fill = self.purple_fill

                    # Add thick border to different cells
                    if key in different_cells:
                        for col in different_cells[key]:
                            # Mark both source and target cells that are different
                            source_cell = ws.cell(row=row_idx + 2,
                                                  column=source_cols[col] + 1)
                            target_cell = ws.cell(row=row_idx + 2,
                                                  column=target_cols[col] + 1)
                            source_cell.border = self.thick_border
                            target_cell.border = self.thick_border

            # Create statistics sheet
            self.create_statistics_sheet(workbook, stats)

            # Auto-adjust column widths
            for worksheet in workbook.worksheets:
                for column in worksheet.columns:
                    max_length = 0
                    column = list(column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

            # Save the file
            writer.close()

            messagebox.showinfo("Siker", f"Az összehasonlítás eredménye mentve:\n{save_path}")

        except Exception as e:
            messagebox.showerror("Hiba", f"Hiba történt az összehasonlítás során:\n{str(e)}")

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    app = FileComparisonApp()
    app.run()
